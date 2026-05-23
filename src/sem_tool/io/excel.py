"""Read and write Excel workbooks for SEM analyses."""

from __future__ import annotations

from pathlib import Path
from typing import Any, Optional

import pandas as pd
from openpyxl import Workbook, load_workbook

from sem_tool.io import schema as sch
from sem_tool.io.diagram import (
    SHEET_DIAGRAMA_SEM,
    data_summary_from_diagram,
    diagrama_sem_template,
)
from sem_tool.io.legend import tipos_variables_dataframe
from sem_tool.io.parameters_rules import (
    SHEET_REGLAS_PARAMETROS,
    ejemplo_hipotesis_v1_v2_template,
    ejemplo_matriz_2x2_template,
    grados_libertad_explicacion,
    identificacion_checklist,
    reglas_parametros_dataframe,
    sistema_3_ecuaciones_template,
    slide_cfa_4_items_template,
)
from sem_tool.cbsem.fit_criteria import fit_criteria_catalog
from sem_tool.io.frederic_curriculum import (
    equivalencias_eqs_smartpls_output,
    identificacion_modelo_frederic,
    indice_taller_frederic,
    leyes_covarianzas_frederic,
    logica_ajuste_chi_cuadrado,
    modelo_uic_calidad_satisfaccion,
)
from sem_tool.plsem.smartpls_guide import (
    modelo_pls_columnas,
    smartpls_vs_cbsem,
    sistema_smartpls_catalog,
)


def read_data(workbook_path: Path) -> pd.DataFrame:
    """Load microdata from sheet Datos."""
    df = pd.read_excel(workbook_path, sheet_name=sch.SHEET_DATOS)
    df = df.dropna(how="all")
    if df.empty:
        raise ValueError(
            f"Hoja '{sch.SHEET_DATOS}' está vacía. "
            "Opciones: (1) Rellenar Datos con sus ítems (CAL1, SAT1, …); "
            "(2) Recrear plantilla con datos de ejemplo: "
            "python3 -m sem_tool init --mode both -o archivo.xlsx --sample"
        )
    return df


def read_covariance(workbook_path: Path) -> Optional[tuple[pd.DataFrame, Optional[pd.Series]]]:
    """Load covariance matrix and optional means (Matriz_Covarianzas / Datos_cov)."""
    cov = None
    for sheet in (sch.SHEET_COVARIANZA, sch.SHEET_DATOS_COV):
        try:
            cov = pd.read_excel(workbook_path, sheet_name=sheet, index_col=0)
            break
        except ValueError:
            continue
    if cov is None:
        return None
    if cov.empty:
        return None
    cov = cov.dropna(how="all").dropna(axis=1, how="all")
    means = None
    try:
        wb = load_workbook(workbook_path, read_only=True, data_only=True)
        if sch.SHEET_MEDIAS in wb.sheetnames:
            means = pd.read_excel(
                workbook_path, sheet_name=sch.SHEET_MEDIAS, index_col=0
            ).squeeze()
        wb.close()
    except Exception:
        pass
    return cov, means


def read_sheet(workbook_path: Path, sheet: str) -> pd.DataFrame:
    df = pd.read_excel(workbook_path, sheet_name=sheet)
    return df.dropna(how="all")


def read_config_value(workbook_path: Path, key: str, default: Any) -> Any:
    try:
        cfg = pd.read_excel(workbook_path, sheet_name=sch.SHEET_CONFIG)
    except ValueError:
        return default
    if cfg.empty or "clave" not in cfg.columns:
        return default
    row = cfg.loc[cfg["clave"].astype(str).str.lower() == key.lower()]
    if row.empty:
        return default
    return row.iloc[0].get("valor", default)


def write_result_sheets(
    workbook_path: Path,
    sheets: dict[str, pd.DataFrame],
    preserve_sheets: Optional[set[str]] = None,
) -> None:
    """Overwrite result sheets; create workbook if missing."""
    path = Path(workbook_path)
    preserve = preserve_sheets or set()

    if path.exists():
        book = load_workbook(path)
        existing = set(book.sheetnames)
        book.close()
        with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            for name, df in sheets.items():
                df.to_excel(writer, sheet_name=name, index=_sheet_needs_index(name))
    else:
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            for name, df in sheets.items():
                df.to_excel(writer, sheet_name=name, index=_sheet_needs_index(name))


def _sheet_needs_index(name: str) -> bool:
    return name in (
        "Fornell_Larcker",
        "HTMT",
        "Paths",
        "R2",
        sch.SHEET_COVARIANZA,
        sch.SHEET_CORRELACION,
        sch.SHEET_MEDIAS,
        sch.SHEET_DATOS_COV,
        sch.SHEET_COVARIANZAS_PARES,
        "Matriz_Observada_DATA",
        "Matriz_Implicita_MODEL",
        "Matriz_Residual",
        "Residuales_Pares",
    )


def create_template(
    output_path: Path,
    mode: str,
    include_sample: bool = True,
) -> None:
    """Create Excel template; por defecto incluye 200 casos simulados (Calidad → Satisfacción)."""
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    from sem_tool.io.sample_data import (
        frederic_datos,
        frederic_modelo_cb,
        frederic_modelo_pls,
        frederic_indicadores,
    )

    # Datos
    ws = wb.create_sheet(sch.SHEET_DATOS)
    if include_sample:
        datos = frederic_datos(n=200)
        ws.append(list(datos.columns))
        for row in datos.itertuples(index=False, name=None):
            ws.append(list(row))
    else:
        ws.append(["CAL1", "CAL2", "CAL3", "SAT1", "SAT2", "SAT3"])
        ws.append([])
        ws.append(["← Pegue aquí sus casos (mínimo 100 filas recomendado)"])

    if mode in ("cb", "both"):
        ws_cb = wb.create_sheet(sch.SHEET_MODELO_CB)
        if include_sample:
            mcb = frederic_modelo_cb()
            ws_cb.append(list(mcb.columns))
            for row in mcb.itertuples(index=False, name=None):
                ws_cb.append(list(row))
        else:
            ws_cb.append(list(sch.CB_COLUMNS))
            ws_cb.append(["medicion", "Calidad", "MEAS", "CAL1 + CAL2 + CAL3", "", "1"])
            ws_cb.append(["medicion", "Satisfaccion", "MEAS", "SAT1 + SAT2 + SAT3", "", "1"])
            ws_cb.append(["estructural", "Satisfaccion", "REG", "Calidad", "", ""])

    if mode in ("pls", "both"):
        ws_spls = wb.create_sheet("Sistema_SmartPLS")
        for title, df in (
            ("Flujo", sistema_smartpls_catalog()),
            ("PLS_vs_CB", smartpls_vs_cbsem()),
            ("Columnas_Modelo_PLS", modelo_pls_columnas()),
        ):
            ws_spls.append([title])
            ws_spls.append(list(df.columns))
            for _, row in df.iterrows():
                ws_spls.append(list(row))
            ws_spls.append([])
        ws_pls = wb.create_sheet(sch.SHEET_MODELO_PLS)
        if include_sample:
            mpl = frederic_modelo_pls()
            ws_pls.append(list(mpl.columns))
            for row in mpl.itertuples(index=False, name=None):
                ws_pls.append(list(row))
        else:
            ws_pls.append(
                ["constructo", "indicador", "modo", "ruta_origen", "ruta_destino", "notas"]
            )
            ws_pls.append(["Calidad", "CAL1", "A", "", "", ""])
            ws_pls.append(["Calidad", "", "A", "Calidad", "Satisfaccion", "H1"])

    _add_methodology_sheets(wb, include_sample=include_sample)
    _add_legend_sheet(wb)
    _add_diagram_sheet(wb)
    _add_parameters_rules_sheet(wb)
    _add_frederic_curriculum_sheets(wb)
    _add_modern_sem_sheets(wb)

    ws_cfg = wb.create_sheet(sch.SHEET_CONFIG)
    ws_cfg.append(["clave", "valor"])
    ws_cfg.append(["observaciones_minimas", 100])
    ws_cfg.append(["bootstraps", 500 if include_sample else 5000])
    ws_cfg.append(["procesos_bootstrap", 2])

    wb.save(path)


def create_restaurant_survey_workbook(
    output_path: Path,
    *,
    n_respondents: int = 20,
    include_data: bool = True,
    min_observations: int = 15,
) -> None:
    """Plantilla restaurante: calidad percibida → satisfacción (demo n=20)."""
    from sem_tool.io.sample_data import (
        frederic_modelo_cb,
        frederic_modelo_pls,
        restaurant_datos,
        restaurant_hipotesis,
        restaurant_indicadores,
    )

    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    ws_items = wb.create_sheet("Items_Encuesta")
    ws_items.append(
        ["indicador", "constructo", "pregunta", "escala", "puntos"]
    )
    for _, row in restaurant_indicadores().iterrows():
        ws_items.append(
            [
                row["indicador"],
                row["constructo"],
                row["notas"],
                row["escala"],
                row["puntos"],
            ]
        )

    ws = wb.create_sheet(sch.SHEET_DATOS)
    if include_data:
        datos = restaurant_datos(n=n_respondents)
        ws.append(list(datos.columns))
        for row in datos.itertuples(index=False, name=None):
            ws.append(list(row))
    else:
        ws.append(["cliente_id", "CAL1", "CAL2", "CAL3", "SAT1", "SAT2", "SAT3"])
        ws.append([])
        ws.append(
            [
                "← Pegue respuestas (1 fila = 1 cliente). "
                "Likert 1–5. Mínimo 3 ítems por constructo."
            ]
        )

    ws_cb = wb.create_sheet(sch.SHEET_MODELO_CB)
    mcb = frederic_modelo_cb()
    ws_cb.append(list(mcb.columns))
    for row in mcb.itertuples(index=False, name=None):
        ws_cb.append(list(row))

    ws_spls = wb.create_sheet("Sistema_SmartPLS")
    for title, df in (
        ("Flujo", sistema_smartpls_catalog()),
        ("PLS_vs_CB", smartpls_vs_cbsem()),
        ("Columnas_Modelo_PLS", modelo_pls_columnas()),
    ):
        ws_spls.append([title])
        ws_spls.append(list(df.columns))
        for _, row in df.iterrows():
            ws_spls.append(list(row))
        ws_spls.append([])

    ws_pls = wb.create_sheet(sch.SHEET_MODELO_PLS)
    mpl = frederic_modelo_pls()
    ws_pls.append(list(mpl.columns))
    for row in mpl.itertuples(index=False, name=None):
        ws_pls.append(list(row))

    ws_h = wb.create_sheet(sch.SHEET_HIPOTESIS)
    hip = restaurant_hipotesis()
    ws_h.append(list(hip.columns))
    for row in hip.itertuples(index=False, name=None):
        ws_h.append(list(row))

    ws_i = wb.create_sheet(sch.SHEET_INDICADORES)
    ind = restaurant_indicadores()
    ws_i.append(list(ind.columns))
    for row in ind.itertuples(index=False, name=None):
        ws_i.append(list(row))

    _add_legend_sheet(wb)
    _add_diagram_sheet(wb)
    _add_parameters_rules_sheet(wb)
    _add_frederic_curriculum_sheets(wb)
    _add_modern_sem_sheets(wb)

    ws_nota = wb.create_sheet("Nota_Muestra")
    ws_nota.append(["concepto", "valor", "comentario"])
    ws_nota.append(
        [
            "casos",
            n_respondents if include_data else 0,
            "Ejemplo pedagógico; para tesis use N>=100",
        ]
    )
    ws_nota.append(
        [
            "modelo",
            "Calidad → Satisfaccion",
            "CB-SEM y PLS-SEM en sem-tool",
        ]
    )

    ws_cfg = wb.create_sheet(sch.SHEET_CONFIG)
    ws_cfg.append(["clave", "valor"])
    ws_cfg.append(["observaciones_minimas", min_observations])
    ws_cfg.append(["bootstraps", 200])
    ws_cfg.append(["procesos_bootstrap", 2])
    ws_cfg.append(
        [
            "contexto",
            "Encuesta restaurante — calidad percibida y satisfacción",
        ]
    )

    wb.save(path)


def _add_methodology_sheets(wb: Workbook, include_sample: bool = True) -> None:
    """Hipótesis fundamentada y catálogo de ítems (literatura + escala)."""
    ws_h = wb.create_sheet(sch.SHEET_HIPOTESIS)
    ws_h.append(["origen", "destino", "hipotesis", "referencia", "argumento"])
    ws_h.append(
        [
            "Calidad",
            "Satisfaccion",
            "H1: La calidad percibida impacta positivamente la satisfacción",
            "Autor (Año) — adaptar cita",
            "Sintetizar hallazgos previos sobre servicio/calidad",
        ]
    )

    ws_i = wb.create_sheet(sch.SHEET_INDICADORES)
    ws_i.append(
        ["constructo", "indicador", "escala", "puntos", "referencia", "notas"]
    )
    for construct, prefix in [("Calidad", "CAL"), ("Satisfaccion", "SAT")]:
        for n in range(1, 4):
            ws_i.append(
                [
                    construct,
                    f"{prefix}{n}",
                    "Likert",
                    5,
                    "Instrumento validado (citar fuente)",
                    "Ordinal; no dicotómica",
                ]
            )


def _add_legend_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet(sch.SHEET_TIPOS_VARIABLES)
    df = tipos_variables_dataframe()
    ws.append(list(df.columns))
    for _, row in df.iterrows():
        ws.append(list(row))


def _add_diagram_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet(SHEET_DIAGRAMA_SEM)
    df = diagrama_sem_template()
    ws.append(list(df.columns))
    for _, row in df.iterrows():
        ws.append(list(row))
    ws2 = wb.create_sheet("Resumen_Diagrama")
    df2 = data_summary_from_diagram()
    ws2.append(list(df2.columns))
    for _, row in df2.iterrows():
        ws2.append(list(row))


def _add_parameters_rules_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet(SHEET_REGLAS_PARAMETROS)
    df = reglas_parametros_dataframe()
    ws.append(list(df.columns))
    for _, row in df.iterrows():
        ws.append(list(row))
    ws2 = wb.create_sheet("Checklist_Identificacion")
    df2 = identificacion_checklist()
    ws2.append(list(df2.columns))
    for _, row in df2.iterrows():
        ws2.append(list(row))
    ws3 = wb.create_sheet("Ejemplo_Matriz_2x2")
    df3 = ejemplo_matriz_2x2_template()
    ws3.append(list(df3.columns))
    for _, row in df3.iterrows():
        ws3.append(list(row))
    ws4 = wb.create_sheet("Hipotesis_V1_V2")
    df4b = ejemplo_hipotesis_v1_v2_template()
    ws4.append(list(df4b.columns))
    for _, row in df4b.iterrows():
        ws4.append(list(row))
    ws_fit = wb.create_sheet("Logica_Ajuste_Modelo")
    df_fit = fit_criteria_catalog()
    ws_fit.append(list(df_fit.columns))
    for _, row in df_fit.iterrows():
        ws_fit.append(list(row))
    ws0 = wb.create_sheet("CFA_1_Factor_4_Items")
    df0 = slide_cfa_4_items_template()
    ws0.append(list(df0.columns))
    for _, row in df0.iterrows():
        ws0.append(list(row))
    ws5 = wb.create_sheet("Sistema_3_Ecuaciones")
    df5 = sistema_3_ecuaciones_template()
    ws5.append(list(df5.columns))
    for _, row in df5.iterrows():
        ws5.append(list(row))
    ws4 = wb.create_sheet("Grados_Libertad")
    df4 = grados_libertad_explicacion()
    ws4.append(list(df4.columns))
    for _, row in df4.iterrows():
        ws4.append(list(row))


def _add_frederic_curriculum_sheets(wb: Workbook) -> None:
    """Hojas alineadas al PPT Frederic Marimon / Marta Mas (UIC)."""
    sheets = [
        ("Curso_Frederic_SEM", indice_taller_frederic()),
        ("Leyes_Covarianzas", leyes_covarianzas_frederic()),
        ("Identificacion_Modelo", identificacion_modelo_frederic()),
        ("Modelo_UIC_F1_F2", modelo_uic_calidad_satisfaccion()),
        ("Logica_Chi_Cuadrado", logica_ajuste_chi_cuadrado()),
        ("EQS_vs_SmartPLS", equivalencias_eqs_smartpls_output()),
    ]
    for name, df in sheets:
        ws = wb.create_sheet(name)
        ws.append(list(df.columns))
        for _, row in df.iterrows():
            ws.append(list(row))


def _add_modern_sem_sheets(wb: Workbook) -> None:
    from sem_tool.cbsem.fit_modern import (
        enfoque_recomendado_moderno,
        fit_criteria_modern_catalog,
    )

    for name, df in (
        ("Criterios_Modernos_SEM", fit_criteria_modern_catalog()),
        ("Enfoque_Moderno_CB_PLS", enfoque_recomendado_moderno()),
    ):
        ws = wb.create_sheet(name)
        ws.append(list(df.columns))
        for _, row in df.iterrows():
            ws.append(list(row))


def merge_templates(cb_path: Path, pls_path: Path, out_path: Path) -> None:
    """Combine CB and PLS templates into one workbook."""
    data_cb = pd.read_excel(cb_path, sheet_name=None)
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for name, df in data_cb.items():
            df.to_excel(writer, sheet_name=name, index=False)
